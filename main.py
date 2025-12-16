import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Confirmação por Mesa", layout="wide")
st.title("✅ Confirmação de presença por Mesa (Excel)")

def nome_brasileiro(texto: str) -> str:
    excecoes = {"da", "de", "do", "das", "dos", "e"}
    partes = texto.strip().lower().split()
    if not partes:
        return ""
    return " ".join(p if p in excecoes else p.capitalize() for p in partes)

verde_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

uploaded = st.file_uploader("Envie sua planilha Excel (.xlsx)", type=["xlsx"])


def achar_linha_cabecalho(ws, max_scan=40):
    for r in range(1, min(ws.max_row, max_scan) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower().startswith("mesa"):
                return r
    return None

def ler_mesas(ws, header_row):
    mesas = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip().lower().startswith("mesa"):
            mesas.append((c, v.strip()))
    return mesas

def montar_itens_por_mesa(ws, header_row, mesas):
    out = {mesa_nome: [] for _, mesa_nome in mesas}
    start_row = header_row + 1

    for col, mesa_nome in mesas:
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            texto = str(v).strip()
            if not texto:
                continue

            out[mesa_nome].append({
                "coord": f"R{r}C{col}",
                "row": r,
                "col": col,
                "texto_formatado": nome_brasileiro(texto)
            })
    return out

def gerar_excel_atualizado(original_bytes: bytes, confirmados_coords: set[str]) -> bytes:
    bio = BytesIO(original_bytes)
    wb = load_workbook(bio)
    ws = wb.active

    header_row = achar_linha_cabecalho(ws)
    if header_row is None:
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    mesas = ler_mesas(ws, header_row)
    start_row = header_row + 1

    # Corrige capitalização nas colunas Mesa...
    for col, _ in mesas:
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(r, col).value
            if isinstance(v, str) and v.strip():
                ws.cell(r, col).value = nome_brasileiro(v)

    # Pinta confirmados
    for coord in confirmados_coords:
        r = int(coord.split("C")[0][1:])
        c = int(coord.split("C")[1])
        ws.cell(r, c).fill = verde_fill

    out2 = BytesIO()
    wb.save(out2)
    return out2.getvalue()


if uploaded:
    original_bytes = uploaded.getvalue()
    wb = load_workbook(BytesIO(original_bytes))
    ws = wb.active

    header_row = achar_linha_cabecalho(ws)
    if header_row is None:
        st.error("Não encontrei a linha do cabeçalho com 'Mesa'.")
        st.stop()

    mesas = ler_mesas(ws, header_row)
    if not mesas:
        st.error("Não encontrei colunas com 'Mesa X'.")
        st.stop()

    itens_por_mesa = montar_itens_por_mesa(ws, header_row, mesas)

    if "confirmados" not in st.session_state:
        st.session_state.confirmados = set()

    st.caption("Clique no nome para confirmar (aparece ✅ e muda o estilo). Clique de novo para desfazer.")

    mesa_nomes = [m for _, m in mesas]
    cols = st.columns(3)

    for i, mesa_nome in enumerate(mesa_nomes):
        with cols[i % 3]:
            with st.expander(mesa_nome, expanded=False):
                lista = itens_por_mesa[mesa_nome]
                if not lista:
                    st.write("—")

                for item in lista:
                    coord = item["coord"]
                    confirmado = coord in st.session_state.confirmados

                    # ✅ feedback visual no próprio texto
                    label = ("✅ " if confirmado else "⬜ ") + item["texto_formatado"]

                    # tenta usar botão "primary" quando confirmado (se sua versão suportar)
                    try:
                        clicked = st.button(
                            label,
                            key=f"btn_{coord}",
                            type="primary" if confirmado else "secondary",
                            use_container_width=True,
                        )
                    except TypeError:
                        # fallback pra versões antigas
                        clicked = st.button(label, key=f"btn_{coord}", use_container_width=True)

                    if clicked:
                        if confirmado:
                            st.session_state.confirmados.discard(coord)
                        else:
                            st.session_state.confirmados.add(coord)
                        st.rerun()

    st.divider()

    excel_bytes = gerar_excel_atualizado(original_bytes, st.session_state.confirmados)

    st.download_button(
        "⬇️ Baixar planilha atualizada (confirmados em verde)",
        data=excel_bytes,
        file_name="confirmacao_atualizada.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.write(f"Confirmados: **{len(st.session_state.confirmados)}**")
else:
    st.info("Envie a planilha .xlsx para começar.")